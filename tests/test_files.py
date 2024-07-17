import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_xlsx_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        # Проверьте, содержится ли test_postback.xlsx в архиве
        assert "test_postback.xlsx" in zip_file.namelist(), "Файл test_postback.xlsx отсутствует в архиве"

        with zip_file.open("test_postback.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=1, column=2).value
            name = 'ID Postback'
            assert name in cell_value, f"Название колонки: {name} есть в файле"


def test_csv_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_postbackcsv.csv") as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csv_reader = list(csv.reader(content.splitlines()))
            second_row = csv_reader[1]
            result_list = second_row
            Postback = "test5"
            IDPostback = "5760956"

            assert result_list[0] == Postback, (f"Название постбэка: {Postback
            } присутствует в таблице {result_list}")
            assert result_list[1] == IDPostback, (f"ID по: {IDPostback
            } присутствует в таблице {IDPostback}")


def test_pdf_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]
            text = page.extract_text()
            assert 'Тестовый PDF файл' in text
