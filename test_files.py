import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_xlsx_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_postback.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=1, column=2).value
            print(f"Содержимое ячейки в строке{1} и столбце {2}: {cell_value}")
            name = 'ID Postback'
            assert name in cell_value, f"Название колонки: {name} есть в файле"